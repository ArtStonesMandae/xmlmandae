import streamlit as st
import pandas as pd
import zipfile
import xml.etree.ElementTree as ET
import openpyxl
from io import BytesIO
from datetime import datetime, timedelta
import re

# Função para normalizar o valor declarado da planilha, tratando vírgulas e garantindo float com 2 casas decimais
def normalizar_valor_planilha(valor_raw):
    if valor_raw is None:
        return None
    if isinstance(valor_raw, (int, float)):
        return round(float(valor_raw), 2)
    try:
        valor_str = str(valor_raw).replace(',', '.')
        return round(float(valor_str), 2)
    except:
        return None

st.title("🔄 Atualizar CHAVE NF com XMLs de Nota Fiscal")
st.write("Atualize a coluna 'CHAVE NF' da planilha Mandae automaticamente com base nos arquivos XML de NF-e.")

planilha_file = st.file_uploader("1. Selecione a planilha Mandae (.xlsx):", type=["xlsx"], key="xlsx_upload")
zip_file = st.file_uploader("2. Selecione o arquivo .ZIP com os XMLs de NF-e:", type=["zip"], key="zip_upload")

if planilha_file and zip_file:
    try:
        wb = openpyxl.load_workbook(planilha_file)
        ws = wb.active

        header = [cell.value for cell in ws[2]]
        idx_doc = header.index("CPF / CNPJ CLIENTE*") + 1
        idx_valor = header.index("VALOR DECLARADO (OPCIONAL)") + 1
        idx_chave = header.index("CHAVE NF") + 1

        regex_cpf = re.compile(r'^\d{11}$')
        regex_cnpj = re.compile(r'^\d{14}$')

        # Lê os XMLs e monta dicionário: (documento, valor) → chave
        doc_valor_para_chave = {}
        with zipfile.ZipFile(zip_file) as z:
            for name in z.namelist():
                if name.endswith(".xml"):
                    with z.open(name) as f:
                        try:
                            tree = ET.parse(f)
                            root = tree.getroot()
                            ns = {'ns': root.tag.split('}')[0].strip('{')}

                            cpf = root.findtext('.//ns:CPF', namespaces=ns)
                            cnpj = root.findtext('.//ns:CNPJ', namespaces=ns)
                            chave = root.findtext('.//ns:chNFe', namespaces=ns)
                            vprod = root.findtext('.//ns:vProd', namespaces=ns)

                            doc = cpf or cnpj
                            if doc and vprod and chave:
                                doc = re.sub(r'\D', '', doc)
                                valor = round(float(vprod), 2)
                                doc_valor_para_chave[(doc, valor)] = chave
                        except:
                            continue

        total_pedidos = 0
        atualizados = 0
        pendencias = []

        # Processa cada linha da planilha
        for row in ws.iter_rows(min_row=3):
            doc_raw = str(row[idx_doc - 1].value).strip()
            valor_raw = row[idx_valor - 1].value

            doc = re.sub(r'\D', '', doc_raw)
            if not (regex_cpf.fullmatch(doc) or regex_cnpj.fullmatch(doc)):
                continue

            valor = normalizar_valor_planilha(valor_raw)
            if valor is None:
                pendencias.append((row[0].row, doc, "Valor ausente ou inválido"))
                continue

            total_pedidos += 1
            chave = doc_valor_para_chave.get((doc, valor))
            if chave:
                ws.cell(row=row[0].row, column=idx_chave, value=chave)
                atualizados += 1
            else:
                pendencias.append((row[0].row, doc, "Chave não encontrada"))

        # Gera a planilha de saída em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Define nome do arquivo com base no próximo dia útil
        hoje = datetime.today()
        dia_util = hoje + timedelta(days=1)
        if hoje.weekday() == 4:
            dia_util += timedelta(days=2)

        nome_arquivo = f"{total_pedidos}Pedidos - {dia_util.strftime('%d.%m')} - L2.xlsx"

        # Mensagens e botões no app
        st.success(f"Chaves atualizadas com sucesso: {atualizados} de {total_pedidos} pedidos.")
        if pendencias:
            st.warning(f"{len(pendencias)} pedidos não puderam ser atualizados. Verifique os dados.")
            with st.expander("🔍 Ver detalhes das pendências"):
                for linha, doc, motivo in pendencias:
                    st.text(f"Linha {linha} | Documento: {doc} | Motivo: {motivo}")

        st.download_button(
            "📅 Baixar Planilha Atualizada",
            data=output,
            file_name=nome_arquivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Erro no processamento: {e}")
